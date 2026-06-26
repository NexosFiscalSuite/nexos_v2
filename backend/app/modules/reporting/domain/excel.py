"""Gera o .xlsx do relatório avançado: abas Capa, Produtos e Alertas.

- Cabeçalho estilizado; colunas financeiras formatadas e somáveis.
- Linha de totais com =SUBTOTAL(9; intervalo) (recalcula com filtros do Excel),
  só nas colunas financeiras (identificadores ficam em branco).
- Colunas de auditoria manuais com fundo amarelo (#FFFACD).
- Aba "Alertas e Inconsistências" só quando há vedações/erros.
"""
from io import BytesIO

_MONEY_FMT = "#,##0.00"
_AUDIT_FILL = "FFFACD"
_HEAD_FILL = "EFEFF1"   # cinza claro (cabeçalho premium, estética minimal)
_HEAD_FONT = "1D1D1F"   # ink (texto do cabeçalho)
_HEAD_BORDER = "D0D0D5"


def _autosize(ws, ncols):
    """Auto-ajuste de largura: dados grandes (ex.: Chave de Acesso, 44 dígitos)
    não ficam esmagados; teto generoso para a chave caber inteira."""
    from openpyxl.utils import get_column_letter
    for i in range(1, ncols + 1):
        maxlen = 12
        for cell in ws[get_column_letter(i)]:
            v = cell.value
            if v is not None:
                maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(maxlen + 3, 52)


def _ref(operand, row, tag2letter, calc2letter, lit):
    """Referência de um operando: célula da tag (onde quer que ela esteja),
    célula de outra coluna calc (@id), ou o número literal entre parênteses
    quando a coluna referenciada não está no relatório."""
    if operand.startswith("@"):
        return f"{calc2letter[operand[1:]]}{row}"
    if operand in tag2letter:
        return f"{tag2letter[operand]}{row}"
    return "(%.2f)" % (lit.get(operand) or 0.0)


def _formula(spec, row, tag2letter, calc2letter, lit):
    def R(op):
        return _ref(op, row, tag2letter, calc2letter, lit)
    if spec["tipo"] == "prod":
        return f"={R(spec['a'])}*{R(spec['b'])}/{spec['div']}"
    parts = []
    for i, (sinal, op) in enumerate(spec["termos"]):
        ref = R(op)
        parts.append((("" if sinal == "+" else "-") + ref) if i == 0 else (sinal + ref))
    return "=" + "".join(parts)


def build_excel(report: dict, nome_modelo: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    audit_fill = PatternFill("solid", fgColor=_AUDIT_FILL)
    head_fill = PatternFill("solid", fgColor=_HEAD_FILL)
    head_font = Font(bold=True, color=_HEAD_FONT)
    head_border = Border(bottom=Side(style="thin", color=_HEAD_BORDER))
    head_align = Alignment(vertical="center")

    wb = Workbook()
    wb.properties.title = nome_modelo

    def _estilizar_cabecalho(ws, ncols, nrows):
        """Cabeçalho premium: negrito + cinza claro, congelado (freeze panes) e
        com auto-filtro para o cliente fatiar os dados (tabelas dinâmicas)."""
        for ci in range(1, ncols + 1):
            cell = ws.cell(row=1, column=ci)
            cell.fill = head_fill
            cell.font = head_font
            cell.border = head_border
            cell.alignment = head_align
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"                   # cabeçalho fixo ao rolar
        if ncols:
            ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(nrows + 1, 1)}"

    def aba(ws, pack, totais):
        cols = pack["cols"]                      # [{label, money, audit?, grupo, calc_spec?}]
        ncols = len(cols)
        ws.append([c["label"] for c in cols])

        # mapa tag/calc -> letra da coluna (posição REAL, p/ as fórmulas)
        tag2letter, calc2letter = {}, {}
        for ci, c in enumerate(cols, start=1):
            L = get_column_letter(ci)
            if c.get("calc_spec"):
                calc2letter[c["calc_id"]] = L
            elif c.get("tag") and c["tag"] not in ("_audit", "_finalidade"):
                tag2letter[c["tag"]] = L

        literals = pack.get("literals") or []
        for ri, datarow in enumerate(pack["rows"]):
            excel_row = ri + 2
            ws.append(list(datarow))  # colunas de dados (calc ficam vazias)
            lit = literals[ri] if ri < len(literals) else {}
            for ci, c in enumerate(cols, start=1):
                if c.get("calc_spec"):
                    ws.cell(row=excel_row, column=ci,
                            value=_formula(c["calc_spec"], excel_row, tag2letter, calc2letter, lit))

        # formatação financeira + fundo das colunas de auditoria (na posição escolhida)
        nrows = len(pack["rows"])
        for ci, c in enumerate(cols, start=1):
            if c["money"]:
                for ri in range(2, nrows + 2):
                    ws.cell(row=ri, column=ci).number_format = _MONEY_FMT
            elif c.get("audit"):
                for ri in range(2, nrows + 2):
                    ws.cell(row=ri, column=ci).fill = audit_fill

        # linha de totais (SUBTOTAL) só nas colunas financeiras
        if totais and nrows:
            trow = nrows + 2
            ws.cell(row=trow, column=1, value="TOTAL").font = Font(bold=True)
            for ci, c in enumerate(cols, start=1):
                if c["money"]:
                    L = get_column_letter(ci)
                    cell = ws.cell(row=trow, column=ci, value=f"=SUBTOTAL(9,{L}2:{L}{nrows + 1})")
                    cell.number_format = _MONEY_FMT
                    cell.font = Font(bold=True)
        _estilizar_cabecalho(ws, ncols, nrows)
        _autosize(ws, ncols)

    ws1 = wb.active
    ws1.title = "Capa"
    aba(ws1, report["capa"], report.get("totais"))

    if report["itens"]["cols"]:
        ws2 = wb.create_sheet("Produtos")
        aba(ws2, report["itens"], report.get("totais"))

    al = report.get("alertas")
    if al:
        ws3 = wb.create_sheet("Alertas e Inconsistências")
        ws3.append(al["cols"])
        warn_fill = PatternFill("solid", fgColor="D9383A")
        for ci in range(1, len(al["cols"]) + 1):
            ws3.cell(row=1, column=ci).font = Font(bold=True, color="FFFFFF")
            ws3.cell(row=1, column=ci).fill = warn_fill
        for r in al["rows"]:
            ws3.append(r)
        ws3.row_dimensions[1].height = 22
        ws3.freeze_panes = "A2"
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(al['cols']))}{len(al['rows']) + 1}"
        _autosize(ws3, len(al["cols"]))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
