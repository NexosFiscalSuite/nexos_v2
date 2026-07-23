"""Planilha Excel das divergências de ICMS-ST (planilha de trabalho do analista).

Aba "Divergências": uma linha por item apontado, com o confronto completo
(declarado × calculado × diferença) e o diagnóstico do motor.
Aba "Memória de cálculo": o passo a passo de CADA item divergente em colunas
agrupadas (alíquotas → MVA → base → apuração → FCP → rastreabilidade) — a
mesma jornada do modal da tela, em formato conferível no Excel.
Aba "Por fornecedor": consolidação para priorizar a cobrança (a mesma régua do
ranking da tela: valor cobrável exclui as antecipações ERRO_111).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NAVY = "24477B"
_NAVY_CLARO = "E8EEF7"
_MONEY_FMT = '#,##0.00;[RED]-#,##0.00'
_PCT_FMT = '#,##0.00"%"'

_DEDUCAO_LEGIVEL = {
    "real": "ICMS da própria nota",
    "teorica": "teórica (emitente do Simples)",
    "zero": "isenta (nada a descontar)",
    "contaminada": "recalculada (própria veio zerada)",
}

_COLUNAS = [
    # (título, chave, largura, formato)
    ("Fornecedor", "fornecedor", 34, None),
    ("CNPJ", "cnpj_emit", 16, None),
    ("NF", "numero_nota", 9, None),
    ("Item", "numero_item", 6, None),
    ("Emissão", "data_emissao", 11, None),
    ("Produto", "descricao", 38, None),
    ("NCM", "ncm", 11, None),
    ("CEST", "cest", 10, None),
    ("CST", "cst_csosn", 6, None),
    ("modBCST", "mod_bc_st", 9, None),
    ("UF O→D", None, 9, None),                       # composta
    ("ST no XML", "vicms_st_xml", 12, _MONEY_FMT),
    ("ST calculado", "vicms_st_calculado", 12, _MONEY_FMT),
    ("Diferença", "diferenca", 12, _MONEY_FMT),
    ("FCP XML", "vfcp_st_xml", 10, _MONEY_FMT),
    ("FCP calc.", "vfcp_st_calculado", 10, _MONEY_FMT),
    ("Status", "status", 15, None),
    ("Código(s) do motor", "codigo_erro", 30, None),
    ("Observação", "observacao", 50, None),
]


def _num(v):
    return float(v) if v is not None else None


def _cabecalho(ws, titulos):
    fill = PatternFill("solid", fgColor=_NAVY)
    fonte = Font(bold=True, color="FFFFFF")
    for col, titulo in enumerate(titulos, start=1):
        c = ws.cell(row=1, column=col, value=titulo)
        c.fill = fill
        c.font = fonte
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}1"


def gerar_xlsx_divergencias(itens: list[dict], titulo_periodo: str) -> bytes:
    wb = Workbook()

    # ── Aba 1: item a item ──
    ws = wb.active
    ws.title = "Divergências"
    _cabecalho(ws, [c[0] for c in _COLUNAS])
    for col, (_t, _k, largura, _f) in enumerate(_COLUNAS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    for r, i in enumerate(itens, start=2):
        for col, (_titulo, chave, _larg, fmt) in enumerate(_COLUNAS, start=1):
            if chave is None:  # UF O→D
                valor = f"{i.get('uf_origem') or '—'}→{i.get('uf_destino') or '—'}"
            elif fmt is _MONEY_FMT:
                valor = _num(i.get(chave))
            else:
                valor = i.get(chave)
            c = ws.cell(row=r, column=col, value=valor)
            if fmt:
                c.number_format = fmt

    _aba_memoria(wb, itens)

    # ── Aba: consolidação por fornecedor (valor cobrável exclui ERRO_111) ──
    ws2 = wb.create_sheet("Por fornecedor")
    _cabecalho(ws2, ["Fornecedor", "CNPJ", "Itens divergentes", "Não auditáveis",
                     "A recolher (fornecedor)", "A favor", "Antecipação (guia própria)"])
    for col, largura in enumerate((34, 16, 16, 14, 20, 14, 22), start=1):
        ws2.column_dimensions[get_column_letter(col)].width = largura

    porf: dict[str, dict] = {}
    for i in itens:
        chave = i.get("cnpj_emit") or "sem-cnpj"
        f = porf.setdefault(chave, {
            "nome": i.get("fornecedor"), "cnpj": i.get("cnpj_emit"),
            "divergentes": 0, "nao_auditaveis": 0,
            "a_recolher": 0.0, "a_favor": 0.0, "antecipacao": 0.0,
        })
        dif = float(i.get("diferenca") or 0)
        antecipa = "ERRO_111" in (i.get("codigo_erro") or "")
        if i.get("status") == "DIVERGENTE":
            f["divergentes"] += 1
            if antecipa:
                f["antecipacao"] += -dif
            elif dif < 0:
                f["a_recolher"] += -dif
            else:
                f["a_favor"] += dif
        else:
            f["nao_auditaveis"] += 1

    ordenado = sorted(porf.values(),
                      key=lambda f: f["a_recolher"] + f["a_favor"], reverse=True)
    for r, f in enumerate(ordenado, start=2):
        ws2.cell(row=r, column=1, value=f["nome"])
        ws2.cell(row=r, column=2, value=f["cnpj"])
        ws2.cell(row=r, column=3, value=f["divergentes"])
        ws2.cell(row=r, column=4, value=f["nao_auditaveis"])
        for col, chave in ((5, "a_recolher"), (6, "a_favor"), (7, "antecipacao")):
            c = ws2.cell(row=r, column=col, value=round(f[chave], 2))
            c.number_format = _MONEY_FMT

    # Rodapé de contexto (auditabilidade do arquivo).
    ws2.cell(row=len(ordenado) + 3, column=1,
             value=f"Nexos Fiscal Suite · Divergências de ICMS-ST · {titulo_periodo}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Bandas da memória: (título do grupo, [(coluna, largura, formato, extrator)]).
def _mf(v):
    """Números da memória chegam como string (JSON) — converte com tolerância."""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


_BANDAS_MEMORIA = [
    ("Identificação", [
        ("Fornecedor", 30, None, lambda i, m: i.get("fornecedor")),
        ("NF / Item", 10, None, lambda i, m: f"{i.get('numero_nota') or '—'} / {i.get('numero_item')}"),
        ("Produto", 34, None, lambda i, m: i.get("descricao")),
        ("NCM", 11, None, lambda i, m: i.get("ncm")),
        ("UF O→D", 9, None, lambda i, m: f"{i.get('uf_origem') or '—'}→{i.get('uf_destino') or '—'}"),
    ]),
    ("Alíquotas", [
        ("Inter", 8, _PCT_FMT, lambda i, m: _mf(m.get("alq_inter"))),
        ("Interna", 9, _PCT_FMT, lambda i, m: _mf(m.get("alq_intra"))),
    ]),
    ("MVA", [
        ("Na nota", 9, _PCT_FMT, lambda i, m: _mf(i.get("pmva_xml"))),
        ("Original", 9, _PCT_FMT, lambda i, m: _mf(m.get("mva_original"))),
        ("Aplicada", 9, _PCT_FMT, lambda i, m: _mf(m.get("mva_aplicada"))),
        ("Ajustada?", 10, None, lambda i, m: "sim" if m.get("mva_foi_ajustada") else "não"),
    ]),
    ("Custo que formou a base", [
        ("Produto", 11, _MONEY_FMT, lambda i, m: _mf(m.get("custo_produto"))),
        ("Frete", 10, _MONEY_FMT, lambda i, m: _mf(m.get("custo_frete"))),
        ("· do CT-e", 10, _MONEY_FMT, lambda i, m: _mf(m.get("custo_frete_cte"))),
        ("IPI", 10, _MONEY_FMT, lambda i, m: _mf(m.get("custo_ipi"))),
        ("(−) Desc.", 10, _MONEY_FMT, lambda i, m: _mf(m.get("custo_desconto"))),
    ]),
    ("Base do ST", [
        ("Na nota", 12, _MONEY_FMT, lambda i, m: _mf(i.get("vbc_st_xml"))),
        ("Calculada", 12, _MONEY_FMT, lambda i, m: _mf(m.get("base_st_calculada"))),
    ]),
    ("Apuração do ICMS-ST", [
        ("Débito (base × alíq.)", 14, _MONEY_FMT, lambda i, m: _mf(m.get("icms_st_debito"))),
        ("(−) ICMS próprio", 13, _MONEY_FMT, lambda i, m: _mf(m.get("deducao_aplicada"))),
        ("Tipo de dedução", 24, None,
         lambda i, m: _DEDUCAO_LEGIVEL.get(m.get("deducao_tipo"), m.get("deducao_tipo"))),
        ("ST devido", 12, _MONEY_FMT, lambda i, m: _mf(m.get("icms_st_calculado"))),
        ("ST na nota", 12, _MONEY_FMT, lambda i, m: _mf(i.get("vicms_st_xml"))),
        ("Diferença", 12, _MONEY_FMT, lambda i, m: _mf(i.get("diferenca"))),
    ]),
    ("FCP-ST", [
        ("Débito", 10, _MONEY_FMT, lambda i, m: _mf(m.get("fcp_st_debito"))),
        ("(−) Próprio", 10, _MONEY_FMT, lambda i, m: _mf(m.get("fcp_st_deducao"))),
        ("Devido", 10, _MONEY_FMT, lambda i, m: _mf(m.get("fcp_st_calculado"))),
        ("Na nota", 10, _MONEY_FMT, lambda i, m: _mf(i.get("vfcp_st_xml"))),
    ]),
    ("Rastreabilidade", [
        ("Regime", 8, None, lambda i, m: m.get("regime")),
        ("Protocolo", 14, None, lambda i, m: (
            "interna (n/a)" if m.get("tem_protocolo") is None
            else ("com acordo" if m.get("tem_protocolo") else "sem acordo")
        )),
        ("Base legal MVA", 24, None, lambda i, m: m.get("mva_base_legal")),
        ("Base legal alíquota", 24, None, lambda i, m: m.get("aliquota_base_legal")),
        ("Motor", 8, None, lambda i, m: m.get("engine_version")),
    ]),
]


def _aba_memoria(wb: Workbook, itens: list[dict]) -> None:
    """Uma linha por item COM memória (pendências sem cálculo ficam de fora):
    cabeçalho em duas linhas — banda do grupo mesclada + coluna."""
    ws = wb.create_sheet("Memória de cálculo")
    fill_banda = PatternFill("solid", fgColor=_NAVY)
    fill_col = PatternFill("solid", fgColor=_NAVY_CLARO)
    fonte_banda = Font(bold=True, color="FFFFFF")
    fonte_col = Font(bold=True, color=_NAVY, size=9)
    centro = Alignment(horizontal="center", vertical="center")

    col = 1
    for banda, colunas in _BANDAS_MEMORIA:
        ini = col
        for titulo, largura, _fmt, _fn in colunas:
            c = ws.cell(row=2, column=col, value=titulo)
            c.fill = fill_col
            c.font = fonte_col
            ws.column_dimensions[get_column_letter(col)].width = largura
            col += 1
        ws.merge_cells(start_row=1, start_column=ini, end_row=1, end_column=col - 1)
        b = ws.cell(row=1, column=ini, value=banda)
        b.fill = fill_banda
        b.font = fonte_banda
        b.alignment = centro

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(col - 1)}2"

    linha = 3
    for i in itens:
        m = i.get("memoria")
        if not m:
            continue                      # NAO_AUDITAVEL: não houve cálculo
        c_idx = 1
        for _banda, colunas in _BANDAS_MEMORIA:
            for _titulo, _larg, fmt, extrator in colunas:
                cel = ws.cell(row=linha, column=c_idx, value=extrator(i, m))
                if fmt:
                    cel.number_format = fmt
                c_idx += 1
        linha += 1
