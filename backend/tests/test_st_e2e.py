"""E2E (ADR-0001 vivo): XML de NF-e + CT-e -> parser -> persistência -> agregação
de frete -> rateio -> motor -> auditoria salva. Prova a autopeça (R$ 177,50).

SQLite async (aiosqlite); FKs não são checadas no SQLite, então criamos só as
tabelas envolvidas. É o ciclo completo da Fase 3 sem precisar de Postgres.
"""
from decimal import Decimal
from uuid import uuid4

import pytest_asyncio
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.database import Base
from app.modules.fiscal.application.auditoria_query import listar_divergencias
from app.modules.fiscal.application.st_audit_service import StAuditService
from app.modules.fiscal.domain.parser import parse_xml
from app.modules.fiscal.infrastructure.matrizes_models import (
    MatrizAliquota,
    MatrizEnquadramentoSt,
    MatrizFcp,
    MatrizMva,
    MatrizProtocoloSt,
)
from app.modules.fiscal.infrastructure.models import (
    AuditoriaIcmsSt,
    NfeCteVinculo,
    Nota,
    NotaItem,
)
from scripts.seed_matrizes import aplicar_seed

# Só as tabelas envolvidas (evita processing_jobs, que usa JSONB do Postgres).
# As FKs para tenants/empresas resolvem no metadata; o SQLite não as fiscaliza.
_TABELAS = [
    MatrizMva.__table__, MatrizEnquadramentoSt.__table__, MatrizFcp.__table__,
    MatrizProtocoloSt.__table__, MatrizAliquota.__table__, Nota.__table__,
    NotaItem.__table__, NfeCteVinculo.__table__, AuditoriaIcmsSt.__table__,
]

_CHAVE_NFE = "1" * 44
_CHAVE_CTE = "3" * 44

# NF-e SP->MG, 2 itens de autopeça (CST 10, modBCST 4). O emitente declarou o
# grupo de ST mas ZEROU vBCST/vICMSST (erro de emissão) -> o motor recalcula.
def _det(nitem: int) -> str:
    return f"""<det nItem="{nitem}">
    <prod><cProd>A{nitem}</cProd><xProd>Autopeca</xProd><NCM>87082919</NCM>
      <CEST>0107500</CEST><CFOP>6404</CFOP><uCom>UN</uCom><qCom>1</qCom>
      <vUnCom>731.35</vUnCom><vProd>731.35</vProd></prod>
    <imposto><ICMS><ICMS10>
      <orig>0</orig><CST>10</CST><vBC>731.35</vBC><pICMS>12.00</pICMS><vICMS>87.76</vICMS>
      <modBCST>4</modBCST><pMVAST>71.78</pMVAST><pICMSST>18.00</pICMSST>
      <vBCST>0.00</vBCST><vICMSST>0.00</vICMSST>
    </ICMS10></ICMS></imposto>
  </det>"""


_NFE = f"""<nfeProc><NFe><infNFe Id="NFe{_CHAVE_NFE}">
  <ide><mod>55</mod><serie>1</serie><nNF>1</nNF><dhEmi>2026-06-01T10:00:00-03:00</dhEmi></ide>
  <emit><CNPJ>11111111000111</CNPJ><xNome>FORNECEDOR SP</xNome><CRT>3</CRT>
    <enderEmit><UF>SP</UF></enderEmit></emit>
  <dest><CNPJ>22222222000122</CNPJ><xNome>CLIENTE MG</xNome>
    <enderDest><UF>MG</UF></enderDest></dest>
  {_det(1)}{_det(2)}
  <total><ICMSTot><vNF>1462.70</vNF></ICMSTot></total>
</infNFe></NFe></nfeProc>"""

# CT-e separado de R$ 136,10 vinculado à NF-e (rateio dá 68,05/item).
_CTE = f"""<cteProc><CTe><infCte Id="CTe{_CHAVE_CTE}">
  <ide><CFOP>6352</CFOP><dhEmi>2026-06-01T10:00:00-03:00</dhEmi><tpCTe>0</tpCTe>
    <nCT>9</nCT><serie>1</serie></ide>
  <emit><CNPJ>33333333000133</CNPJ><xNome>TRANSPORTADORA</xNome>
    <enderEmit><UF>SP</UF></enderEmit></emit>
  <vPrest><vTPrest>136.10</vTPrest></vPrest>
  <infCTeNorm><infDoc><infNFe><chave>{_CHAVE_NFE}</chave></infNFe></infDoc></infCTeNorm>
</infCte></CTe></cteProc>"""

def _d(v) -> Decimal:
    return Decimal(str(v or 0))


@pytest_asyncio.fixture
async def sessao():
    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all, tables=_TABELAS)
    async with async_sessionmaker(engine, class_=AsyncSession)() as s:
        await aplicar_seed(s)
        await s.commit()
        yield s
    await engine.dispose()


async def _injetar(s: AsyncSession, tenant_id, empresa_id):
    """Parser (XML) -> persistência, como faria a importação real."""
    nfe = parse_xml(_NFE.encode("utf-8"))
    cte = parse_xml(_CTE.encode("utf-8"))

    nota = Nota(
        id=uuid4(), tenant_id=tenant_id, empresa_id=empresa_id,
        chave_acesso=nfe["chave_acesso"], tipo="NFe", fluxo="entrada", modelo="55",
        uf_emit=nfe["uf_emit"], uf_dest=nfe["uf_dest"], crt_emit=nfe["crt_emit"],
        nome_emit=nfe["nome_emit"], cnpj_emit=nfe["cnpj_emit"],
        data_emissao=nfe["data_emissao"],
    )
    s.add(nota)
    for it in nfe["itens"]:
        s.add(NotaItem(
            id=uuid4(), tenant_id=tenant_id, nota_id=nota.id,
            numero_item=it["numero_item"], descricao=it.get("descricao"),
            codigo=it.get("codigo"), ncm=it["ncm"], cest=it["cest"], cfop=it["cfop"],
            orig=it["orig"], cst=it["cst"], csosn=it["csosn"] or None, mod_bc_st=it["mod_bc_st"],
            quantidade=_d(it["quantidade"]), valor_produto=_d(it["valor_produto"]),
            valor_frete=_d(it["valor_frete"]), valor_seguro=_d(it["valor_seguro"]),
            valor_outro=_d(it["valor_outro"]), valor_ipi=_d(it["valor_ipi"]),
            base_calculo=_d(it["base_calculo"]), valor_icms=_d(it["valor_icms"]),
            p_icms=_d(it["p_icms"]), p_red_bc=_d(it["p_red_bc"]),
            p_mva_st=_d(it["p_mva_st"]), p_red_bc_st=_d(it["p_red_bc_st"]),
            p_icms_st=_d(it["p_icms_st"]), v_bc_st=_d(it["v_bc_st"]),
            valor_icms_st=_d(it["valor_icms_st"]),
            v_fcp=_d(it["v_fcp"]), p_fcp=_d(it["p_fcp"]), v_bc_fcp=_d(it["v_bc_fcp"]),
            v_fcp_st=_d(it["v_fcp_st"]), p_fcp_st=_d(it["p_fcp_st"]),
            v_bc_fcp_st=_d(it["v_bc_fcp_st"]),
        ))
    for chave_nfe in cte["chaves_nfe"]:
        s.add(NfeCteVinculo(
            id=uuid4(), tenant_id=tenant_id, empresa_id=empresa_id,
            chave_nfe=chave_nfe, chave_cte=cte["chave_acesso"],
            vtprest=_d(cte["vtprest"]), tp_cte=cte["tp_cte"],
        ))
    await s.flush()
    return nota


async def test_e2e_autopeca_frete_agregado_ate_auditoria(sessao):
    tenant_id, empresa_id = uuid4(), uuid4()
    nota = await _injetar(sessao, tenant_id, empresa_id)

    registros = await StAuditService(sessao).auditar_nota(empresa_id, nota.id)

    print("\n" + "=" * 64)
    print(" E2E — Autopeça SP->MG com CT-e separado (ADR-0001)")
    print("=" * 64)
    print("  CT-e agregado (vTPrest)   : R$ 136,10  -> rateio 2 itens = 68,05/item")
    for r in registros:
        print(f"  Item {r.numero_item}: base ST calc={r.vbc_st_calculado} "
              f"(prova frete na base) | MVA {r.pmva_calculada}% | "
              f"ICMS-ST calc={r.vicms_st_calculado} | XML={r.vicms_st_xml} | "
              f"{r.status} ({r.codigo_erro})")
    salvos = await sessao.scalar(select(func.count()).select_from(AuditoriaIcmsSt))
    print(f"  Persistido em auditoria_icms_st: {salvos} registro(s)")
    print("=" * 64)

    assert len(registros) == 2
    for r in registros:
        # A base 1473,69 só fecha SE o frete (68,05) entrou na base do ST.
        assert r.vbc_st_calculado == Decimal("1473.69")
        assert r.pmva_calculada == Decimal("84.35")          # MVA 71,78 -> ajustada
        assert r.vicms_st_calculado == Decimal("177.50")     # gabarito do laboratório
        assert r.vicms_st_xml == Decimal("0.00")             # emitente zerou o ST
        assert r.status == "DIVERGENTE"
        assert "ERRO_104_VALOR_ST_DIVERGENTE" in r.codigo_erro
        assert r.memoria["mva_original"] == "71.78"          # veio do banco
        # Rastreabilidade (defensibilidade): versão do motor + linhas de matriz
        # usadas + fonte da decisão de protocolo, tudo na memória persistida.
        assert r.memoria["engine_version"]
        assert r.memoria["mva_matriz_id"] is not None
        assert r.memoria["aliquota_matriz_id"] is not None   # alíquota veio da matriz
        assert r.memoria["alq_intra"] == "18.00"             # MG vigente em 2026-06-01
        assert r.memoria["tem_protocolo"] is True            # SP→MG (seed)
        assert r.memoria["protocolo_fonte"] == "matriz"

    # Persistiu de fato (não só retornou).
    assert await sessao.scalar(select(func.count()).select_from(AuditoriaIcmsSt)) == 2


async def test_query_divergencias_e_idempotencia(sessao):
    tenant_id, empresa_id = uuid4(), uuid4()
    nota = await _injetar(sessao, tenant_id, empresa_id)
    svc = StAuditService(sessao)
    await svc.auditar_nota(empresa_id, nota.id)
    await svc.auditar_nota(empresa_id, nota.id)   # re-auditoria NÃO duplica

    assert await sessao.scalar(select(func.count()).select_from(AuditoriaIcmsSt)) == 2

    # Endpoint: REL_Divergencia_ST.
    res = await listar_divergencias(sessao, empresa_id=empresa_id)
    assert res["total"] == 2
    item = res["itens"][0]
    assert item["vicms_st_calculado"] == Decimal("177.50")
    assert item["vicms_st_xml"] == Decimal("0.00")
    assert item["diferenca"] == Decimal("-177.50")       # XML − calculado
    assert item["fornecedor"] == "FORNECEDOR SP"
    assert item["descricao"]                             # nome do produto na listagem
    assert item["ncm"]
    assert item["memoria"]["mva_original"] == "71.78"    # memória exposta ao front

    # Agregados do período: os cards da tela (dinheiro em jogo) + ranking.
    assert res["resumo"]["a_recolher"] == 355.0          # 2 × 177,50 retidos a menor
    assert res["resumo"]["a_favor"] == 0.0
    assert res["resumo"]["antecipacao"] == 0.0
    assert res["resumo"]["divergentes"] == 2
    top = res["ranking_fornecedores"][0]
    assert top["nome"] == "FORNECEDOR SP" and top["itens"] == 2 and top["valor"] == 355.0

    # Filtro de período exclui notas fora da janela.
    vazio = await listar_divergencias(sessao, empresa_id=empresa_id, data_inicio="2027-01-01")
    assert vazio["total"] == 0
    assert vazio["resumo"]["divergentes"] == 0 and vazio["ranking_fornecedores"] == []

    # Filtros de servidor: status, código do motor e busca livre.
    so_div = await listar_divergencias(sessao, empresa_id=empresa_id, status="DIVERGENTE")
    assert so_div["total"] == 2
    nada = await listar_divergencias(sessao, empresa_id=empresa_id, status="NAO_AUDITAVEL")
    assert nada["total"] == 0
    por_cod = await listar_divergencias(sessao, empresa_id=empresa_id, codigo_erro="ERRO_104")
    assert por_cod["total"] == 2
    por_q = await listar_divergencias(sessao, empresa_id=empresa_id, q="autopeca")
    assert por_q["total"] == 2                            # descrição do item (sem caixa)
    por_q2 = await listar_divergencias(sessao, empresa_id=empresa_id, q="fornecedor sp")
    assert por_q2["total"] == 2                           # nome do emitente
    por_q3 = await listar_divergencias(sessao, empresa_id=empresa_id, q="produto inexistente")
    assert por_q3["total"] == 0


async def test_diagnostico_executivo_e_pdf(sessao):
    """Diagnóstico do período: agregados por competência (inclui itens OK),
    top fornecedores e o PDF timbrado do relatório executivo."""
    from app.modules.fiscal.application.auditoria_query import diagnostico_st
    from app.modules.fiscal.application.st_diagnostico import gerar_diagnostico_pdf

    tenant_id, empresa_id = uuid4(), uuid4()
    nota = await _injetar(sessao, tenant_id, empresa_id)
    await StAuditService(sessao).auditar_nota(empresa_id, nota.id)

    d = await diagnostico_st(sessao, empresa_id=empresa_id)
    assert [c["competencia"] for c in d["competencias"]] == ["2026-06"]
    c = d["competencias"][0]
    assert c["itens"] == 2 and c["divergentes"] == 2 and c["ok"] == 0
    assert c["a_recolher"] == 355.0
    assert d["totais"]["pct_conformidade"] == 0.0
    assert d["top_fornecedores"][0]["nome"] == "FORNECEDOR SP"

    pdf = gerar_diagnostico_pdf(
        empresa_nome="CLIENTE MG LTDA", empresa_cnpj="22222222000122", dados=d
    )
    assert pdf[:5] == b"%PDF-" and len(pdf) > 3000

    # Fora do período: 404 amigável fica no router; aqui, estrutura vazia.
    vazio = await diagnostico_st(sessao, empresa_id=empresa_id, data_inicio="2027-01-01")
    assert vazio["competencias"] == [] and vazio["totais"]["itens"] == 0


async def test_exportacao_xlsx_das_divergencias(sessao):
    """Planilha de trabalho: todas as linhas do filtro + consolidação por
    fornecedor, legível pelo openpyxl (o que o Excel abre, o teste abre)."""
    import io

    from openpyxl import load_workbook

    from app.modules.fiscal.application.auditoria_query import exportar_divergencias
    from app.modules.fiscal.application.st_export import gerar_xlsx_divergencias

    tenant_id, empresa_id = uuid4(), uuid4()
    nota = await _injetar(sessao, tenant_id, empresa_id)
    await StAuditService(sessao).auditar_nota(empresa_id, nota.id)

    itens = await exportar_divergencias(sessao, empresa_id=empresa_id)
    assert len(itens) == 2
    xlsx = gerar_xlsx_divergencias(itens, "06/2026")

    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["Divergências", "Memória de cálculo", "Por fornecedor"]
    ws = wb["Divergências"]
    assert ws.cell(row=1, column=1).value == "Fornecedor"
    assert ws.cell(row=2, column=1).value == "FORNECEDOR SP"
    assert ws.cell(row=2, column=6).value == "Autopeca"          # produto
    assert float(ws.cell(row=2, column=13).value) == 177.50      # ST calculado

    # Memória: o passo a passo em colunas (linhas 1-2 = banda + coluna).
    wm = wb["Memória de cálculo"]
    assert wm.cell(row=1, column=1).value == "Identificação"
    cabecalhos = [wm.cell(row=2, column=c).value for c in range(1, 34)]
    assert "Aplicada" in cabecalhos and "ST devido" in cabecalhos
    col_devido = cabecalhos.index("ST devido") + 1
    assert float(wm.cell(row=3, column=col_devido).value) == 177.50
    col_ded = cabecalhos.index("Tipo de dedução") + 1
    assert "nota" in wm.cell(row=3, column=col_ded).value        # "ICMS da própria nota"

    wf = wb["Por fornecedor"]
    assert wf.cell(row=2, column=1).value == "FORNECEDOR SP"
    assert float(wf.cell(row=2, column=5).value) == 355.0        # a recolher
