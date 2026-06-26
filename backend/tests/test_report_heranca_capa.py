"""Relatório de ITENS: tags de capa (Identificação/Emitente) herdadas e repetidas
em cada linha de produto (injeção capa→item)."""
import io
from types import SimpleNamespace

from openpyxl import load_workbook

from app.modules.reporting.domain.excel import build_excel
from app.modules.reporting.domain.report_gen import gerar

_CHAVE = "35260112345678901234567890123456789012345678"
_XML = f"""<nfeProc>
  <NFe><infNFe Id="NFe{_CHAVE}">
    <ide><nNF>55</nNF><natOp>Venda</natOp></ide>
    <emit><CNPJ>11111111000111</CNPJ><xNome>FORNECEDOR SP</xNome></emit>
    <det nItem="1"><prod><xProd>Produto A</xProd></prod></det>
    <det nItem="2"><prod><xProd>Produto B</xProd></prod></det>
  </infNFe></NFe>
  <protNFe><infProt><chNFe>{_CHAVE}</chNFe></infProt></protNFe>
</nfeProc>""".encode()


def test_tags_de_capa_sao_repetidas_em_cada_linha_de_item():
    nota = SimpleNamespace(numero="55", serie="1", chave_acesso=_CHAVE, nome_emit="FORNECEDOR SP")
    config = {
        "capa": [],
        "itens": [{"tag": "chNFe"}, {"tag": "emit_CNPJ"}, {"tag": "it_xProd"}],
        "finalidade": False, "calculos": False, "totais": False,
    }
    res = gerar(config, [{"nota": nota, "tipos": {1: "", 2: ""}, "xml": _XML}], regime="Simples Nacional")

    itens = res["itens"]
    grupos = [c["grupo"] for c in itens["cols"]]
    assert "Identificação" in grupos and "Emitente" in grupos   # tags de capa entraram nos itens
    assert len(itens["rows"]) == 2

    # A chave de acesso e o CNPJ do emitente repetem em TODAS as linhas de produto.
    for linha in itens["rows"]:
        assert _CHAVE in linha
        assert "11111111000111" in linha
    assert "Produto A" in itens["rows"][0]
    assert "Produto B" in itens["rows"][1]


def test_excel_premium_freeze_cabecalho_e_autofilter():
    nota = SimpleNamespace(numero="55", serie="1", chave_acesso=_CHAVE, nome_emit="FORN")
    config = {
        "capa": [{"tag": "nNF"}],
        "itens": [{"tag": "chNFe"}, {"tag": "it_xProd"}],
        "finalidade": False, "calculos": False, "totais": False,
    }
    rep = gerar(config, [{"nota": nota, "tipos": {1: "", 2: ""}, "xml": _XML}], regime="Simples Nacional")
    wb = load_workbook(io.BytesIO(build_excel(rep, "Modelo Oficial")))
    ws = wb["Produtos"]

    assert ws.freeze_panes == "A2"                     # cabeçalho congelado
    assert ws.auto_filter.ref.startswith("A1:")        # filtro p/ tabela dinâmica
    head = ws.cell(row=1, column=1)
    assert head.font.bold is True
    assert head.fill.fgColor.rgb.endswith("EFEFF1")    # cinza claro
    assert ws.column_dimensions["A"].width >= 44       # Chave de Acesso não esmagada


def test_templates_de_fabrica_config_valida():
    from app.modules.reporting.application.service import _TEMPLATES, _valida_config
    from app.modules.reporting.domain.tags import TAGS

    assert {t["nome"] for t in _TEMPLATES} == {"Apuração de ICMS-ST", "Conferência de Entradas"}
    assert "it_pMVAST" in TAGS                          # tag de MVA adicionada
    for t in _TEMPLATES:
        cfg = _valida_config(t["config"])               # não levanta DomainError
        for col in [*cfg["capa"], *cfg["itens"]]:
            assert col["tag"] in TAGS                    # toda tag do template existe
